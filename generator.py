from __future__ import annotations

import re
import unicodedata
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class GenerationError(Exception):
    """Raised when an uploaded workbook cannot be transformed."""


@dataclass(frozen=True)
class GenerationMetadata:
    study_name: str
    protocol_code: str


def generate_field_book(
    source_path: Path,
    template_path: Path,
    output_path: Path,
    repetitions: int = 4,
) -> GenerationMetadata:
    if not template_path.exists():
        raise GenerationError("No se encontró la plantilla del Libro de Campo en el repo.")

    source_wb = load_workbook(source_path, data_only=False)
    if len(source_wb.worksheets) != 4:
        raise GenerationError("El Excel fuente debe tener exactamente 4 hojas.")

    template_wb = load_workbook(template_path, keep_vba=False)
    study_title = _cell_text(source_wb.worksheets[0]["A4"].value)
    protocol_raw = _cell_text(source_wb.worksheets[0]["A6"].value)
    protocol_code = _extract_protocol_code(protocol_raw) or _safe_title(source_path.stem)
    study_name = _safe_title(study_title or protocol_code)

    _build_plano(template_wb["Plano"], study_title or protocol_code)
    _build_aplicaciones(template_wb["Aplicaciones"], source_wb.worksheets[2])
    _build_tratamientos(template_wb["Tratamientos"], source_wb.worksheets[3])
    _build_datos_a_completar(
        template_wb["Datos a completar"],
        source_wb.worksheets[1],
        source_wb.worksheets[3],
        repetitions,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    template_wb.save(output_path)
    return GenerationMetadata(study_name=study_name, protocol_code=protocol_code)


def _build_plano(sheet: Worksheet, title: str) -> None:
    sheet["A1"] = title


def _build_aplicaciones(output: Worksheet, source: Worksheet) -> None:
    table = _extract_application_table(source)
    app_count = max(max(len(row) for row in table) - 1, 1)
    equipment_start_col = app_count + 3
    max_col = equipment_start_col + app_count
    max_row = max(output.max_row, len(table), 34)

    _unmerge_overlapping(output, min_row=1, min_col=1, max_row=max_row, max_col=max(output.max_column, max_col))
    _clear_range(output, min_row=1, min_col=2, max_row=max_row, max_col=max(output.max_column, max_col))

    for col_idx in range(2, 2 + app_count):
        output.cell(2, col_idx, f"Aplicación {_excel_app_letter(col_idx - 1)}")
        for row_idx in range(1, 35):
            _copy_style(output.cell(row_idx, 2), output.cell(row_idx, col_idx))
        output.column_dimensions[get_column_letter(col_idx)].width = 18

    formula_cols = range(2, 2 + app_count)
    for col_idx in formula_cols:
        letter = get_column_letter(col_idx)
        output.cell(23, col_idx, f"={letter}21*{letter}20*{letter}24/60000")

    for row_idx, row in enumerate(table, start=1):
        for col_idx, value in enumerate(row, start=equipment_start_col):
            cell = output.cell(row=row_idx, column=col_idx, value=value)
            style_col = 4 if col_idx == equipment_start_col else 5
            _copy_style(output.cell(row=min(row_idx, output.max_row), column=style_col), cell)

    _style_application_range(
        output,
        rows=len(table),
        cols=max(len(r) for r in table),
        start_col=equipment_start_col,
    )
    _set_range_border(output, 1, 1, 34, max(2 + app_count - 1, 2), outside_style="thick")
    _set_range_border(output, 1, equipment_start_col, len(table), max_col, outside_style="thick")
    output.freeze_panes = "B3"


def _build_tratamientos(output: Worksheet, source: Worksheet) -> None:
    table = _extract_treatment_table(source)
    _clear_range(output, min_row=1, min_col=1, max_row=max(output.max_row, len(table) + 5), max_col=max(output.max_column, len(table[0]) + 2))

    for row_idx, row in enumerate(table, start=1):
        style_source_row = 1 if row_idx <= 2 else min(row_idx, 3)
        for col_idx, value in enumerate(row, start=1):
            cell = output.cell(row=row_idx, column=col_idx, value=value)
            _copy_style(output.cell(row=style_source_row, column=min(col_idx, 9)), cell)

    max_col = max(len(row) for row in table)
    header = output.iter_rows(min_row=1, max_row=2, min_col=1, max_col=max_col)
    for row in header:
        for cell in row:
            cell.fill = copy(output["A1"].fill)
            cell.font = copy(output["A1"].font)
            cell.alignment = copy(output["A1"].alignment)

    _merge_treatment_numbers(output, first_data_row=3, last_row=len(table))
    _autofit_columns(output, 1, max_col, 1, len(table))
    _set_range_border(output, 1, 1, len(table), max_col, outside_style="thick")
    output.freeze_panes = "A3"


def _build_datos_a_completar(output: Worksheet, source: Worksheet, title: str, protocol_code: str) -> None:
    max_source_col = _last_non_empty_col(source, row=10)
    if max_source_col < 2:
        raise GenerationError("No encontré columnas de evaluación en la segunda hoja.")

    data_cols = list(range(2, max_source_col + 1))
    output_max_col = 4 + len(data_cols)
    _clear_range(output, min_row=1, min_col=1, max_row=max(output.max_row, 11), max_col=max(output.max_column, output_max_col))

    rows = [
        [title] + [None] * (output_max_col - 1),
        [protocol_code, None, None, None] + [_evaluation_day(source.cell(23, c).value) for c in data_cols],
        ["Evaluación", None, None, None] + [_evaluation_label(source, c) for c in data_cols],
        ["Unidad", None, None, None] + [_unit_from_text(source.cell(15, c).value) for c in data_cols],
        ["Parte a evaluar", None, None, None] + [_clean_code(source.cell(13, c).value) for c in data_cols],
        ["Cultivo", None, None, None] + [_clean_code(source.cell(20, c).value) for c in data_cols],
        ["Target", None, None, None] + [_clean_code(source.cell(22, c).value) for c in data_cols],
        ["Momento de evaluación", None, None, None] + [_clean_code(source.cell(23, c).value) for c in data_cols],
        ["Fenología", None, None, None] + [""] * len(data_cols),
        ["Fecha", None, None, None] + [""] * len(data_cols),
        ["Repetición", "Parcela", "Tratamiento", "Submuestra"] + [""] * len(data_cols),
    ]

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = output.cell(row=row_idx, column=col_idx, value=value)
            style_col = min(col_idx, 5 if col_idx > 4 else col_idx)
            _copy_style(output.cell(row=min(row_idx, 11), column=style_col), cell)

    for col_idx in range(1, output_max_col + 1):
        output.column_dimensions[get_column_letter(col_idx)].width = 13

    _apply_table_borders(output, 1, 1, 11, output_max_col)
    output.freeze_panes = "E11"


def _extract_application_table(sheet: Worksheet) -> list[list[object]]:
    start_row = None
    for row in range(1, sheet.max_row - 2):
        current = _norm(sheet.cell(row, 1).value)
        next_one = _norm(sheet.cell(row + 1, 1).value)
        next_two = _norm(sheet.cell(row + 2, 1).value)
        if current == "equipo de aplicacion" and not next_one and next_two == "momento de aplicacion":
            start_row = row
            break

    if start_row is None:
        for row in range(1, sheet.max_row + 1):
            if _norm(sheet.cell(row, 1).value) == "momento de aplicacion":
                start_row = max(1, row - 1)
                break

    if start_row is None:
        raise GenerationError("No encontré la tabla de aplicaciones en la tercera hoja.")

    max_col = _last_non_empty_col_in_rows(sheet, range(start_row, min(sheet.max_row, start_row + 13) + 1))
    table = []
    for row in range(start_row, min(sheet.max_row, start_row + 13) + 1):
        table.append([sheet.cell(row, col).value for col in range(1, max_col + 1)])
    return table


def _extract_treatment_table(sheet: Worksheet) -> list[list[object]]:
    header_row = None
    for row in range(1, sheet.max_row):
        if _norm(sheet.cell(row, 1).value) in {"no", "n"} and _norm(sheet.cell(row + 1, 1).value) in {"tr", "trat"}:
            header_row = row
            break

    if header_row is None:
        raise GenerationError("No encontré el inicio de la tabla de tratamientos en la cuarta hoja.")

    max_col = _last_non_empty_col(sheet, header_row)
    last_row = header_row + 1
    for row in range(header_row + 2, sheet.max_row + 1):
        product = sheet.cell(row, 2).value
        if product is None or str(product).strip() == "":
            break
        last_row = row

    table = []
    for row in range(header_row, last_row + 1):
        table.append([sheet.cell(row, col).value for col in range(1, max_col + 1)])
    return table


def _style_application_range(sheet: Worksheet, rows: int, cols: int, start_col: int = 4) -> None:
    max_col = start_col + cols - 1
    for row in range(1, rows + 1):
        for col in range(start_col, max_col + 1):
            cell = sheet.cell(row, col)
            cell.alignment = copy(sheet["E3"].alignment)
            if row == 1:
                cell.fill = copy(sheet["D1"].fill)
                cell.font = copy(sheet["D1"].font)
            elif row == 2:
                cell.fill = copy(sheet["E2"].fill)
                cell.font = copy(sheet["E2"].font)
            else:
                cell.fill = copy(sheet["E3"].fill)
                cell.font = copy(sheet["E3"].font)
            sheet.column_dimensions[get_column_letter(col)].width = 22 if col == start_col else 18
    _set_range_border(sheet, 1, start_col, rows, max_col, outside_style="thick")


def _build_datos_a_completar(
    output: Worksheet,
    evaluation_source: Worksheet,
    treatment_source: Worksheet,
    repetitions: int,
) -> None:
    rows_to_keep = [3, 11, 12, 13, 15, 16, 19, 20, 22, 23]
    data_cols = [
        col
        for col in range(2, evaluation_source.max_column + 1)
        if evaluation_source.cell(10, col).value not in (None, "")
        and not _has_brackets(evaluation_source.cell(25, col).value)
    ]
    if not data_cols:
        raise GenerationError("No encontré columnas válidas en la hoja 2 con la fila 25 vacía.")

    max_subsamples = max(_max_numeric_value(evaluation_source.cell(19, col).value for col in data_cols), 1)
    treatment_count = _count_treatments(treatment_source)
    if treatment_count < 1:
        raise GenerationError("No encontré tratamientos en la cuarta hoja.")

    header_rows = len(rows_to_keep) + 2
    data_header_row = header_rows + 1
    data_rows = repetitions * treatment_count * max_subsamples
    output_max_col = 4 + len(data_cols)
    output_max_row = data_header_row + data_rows

    _unmerge_overlapping(
        output,
        min_row=1,
        min_col=1,
        max_row=max(output.max_row, output_max_row),
        max_col=max(output.max_column, output_max_col),
    )
    _clear_range(
        output,
        min_row=1,
        min_col=1,
        max_row=max(output.max_row, output_max_row),
        max_col=max(output.max_column, output_max_col),
    )

    for out_row, src_row in enumerate(rows_to_keep, start=1):
        output.merge_cells(start_row=out_row, start_column=1, end_row=out_row, end_column=4)
        output.cell(out_row, 1, evaluation_source.cell(src_row, 1).value)
        for out_col, src_col in enumerate(data_cols, start=5):
            output.cell(out_row, out_col, evaluation_source.cell(src_row, src_col).value)

    for row_idx, label in [(header_rows - 1, "Fenología"), (header_rows, "Fecha")]:
        output.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        output.cell(row_idx, 1, label)

    for col, header in enumerate(["Repetición", "Parcela", "Tratamiento", "Submuestra"], start=1):
        output.cell(data_header_row, col, header)

    row_cursor = data_header_row + 1
    for repetition in range(1, repetitions + 1):
        for treatment in range(1, treatment_count + 1):
            for subsample in range(1, max_subsamples + 1):
                output.cell(row_cursor, 1, repetition)
                output.cell(row_cursor, 2, None)
                output.cell(row_cursor, 3, treatment)
                output.cell(row_cursor, 4, subsample)
                row_cursor += 1

    _style_datos_a_completar(output, data_cols, header_rows, data_header_row, output_max_col, output_max_row)
    if output.max_column > output_max_col:
        output.delete_cols(output_max_col + 1, output.max_column - output_max_col)
    output.freeze_panes = f"E{data_header_row + 1}"
    output.auto_filter.ref = f"A{data_header_row}:{get_column_letter(output_max_col)}{output_max_row}"


def _clear_range(sheet: Worksheet, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            sheet.cell(row=row, column=col).value = None


def _unmerge_overlapping(sheet: Worksheet, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    ranges = list(sheet.merged_cells.ranges)
    for merged in ranges:
        if (
            merged.max_row >= min_row
            and merged.min_row <= max_row
            and merged.max_col >= min_col
            and merged.min_col <= max_col
        ):
            sheet.unmerge_cells(str(merged))


def _copy_style(source, target) -> None:
    if source.has_style:
        try:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
        except AttributeError:
            return


def _apply_table_borders(sheet: Worksheet, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    border = copy(sheet.cell(min_row, min_col).border)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            try:
                sheet.cell(row, col).border = border
            except AttributeError:
                continue


def _set_range_border(
    sheet: Worksheet,
    min_row: int,
    min_col: int,
    max_row: int,
    max_col: int,
    outside_style: str = "thin",
) -> None:
    thin = Side(style="thin", color="808080")
    outside = Side(style=outside_style, color="404040")
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            try:
                sheet.cell(row, col).border = Border(
                    left=outside if col == min_col else thin,
                    right=outside if col == max_col else thin,
                    top=outside if row == min_row else thin,
                    bottom=outside if row == max_row else thin,
                )
            except AttributeError:
                continue


def _autofit_columns(sheet: Worksheet, min_col: int, max_col: int, min_row: int, max_row: int) -> None:
    for col in range(min_col, max_col + 1):
        max_len = 0
        for row in range(min_row, max_row + 1):
            value = sheet.cell(row, col).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        width = min(max(max_len + 2, 8), 45)
        if col == 1:
            width = min(width, 12)
        sheet.column_dimensions[get_column_letter(col)].width = width


def _merge_treatment_numbers(sheet: Worksheet, first_data_row: int, last_row: int) -> None:
    current_start = None
    for row in range(first_data_row, last_row + 2):
        value = sheet.cell(row, 1).value if row <= last_row else "END"
        if isinstance(value, (int, float)) or value == "END":
            if current_start is not None and row - 1 > current_start:
                sheet.merge_cells(start_row=current_start, start_column=1, end_row=row - 1, end_column=1)
                cell = sheet.cell(current_start, 1)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            current_start = row if value != "END" else None


def _excel_app_letter(index: int) -> str:
    return get_column_letter(index)


def _style_datos_a_completar(
    sheet: Worksheet,
    data_cols: list[int],
    header_rows: int,
    data_header_row: int,
    max_col: int,
    max_row: int,
) -> None:
    label_fill = PatternFill("solid", fgColor="DDEBF7")
    data_header_fill = PatternFill("solid", fgColor="E2EFDA")
    no_fill = PatternFill(fill_type=None)

    for col in range(1, max_col + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 14 if col <= 4 else 16

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            style_row = min(row, 11)
            style_col = min(col, 5 if col > 4 else col)
            _copy_style(sheet.cell(style_row, style_col), sheet.cell(row, col))

    for row in range(1, header_rows + 1):
        for col in range(1, 5):
            try:
                sheet.cell(row, col).fill = copy(label_fill)
                sheet.cell(row, col).font = copy(sheet["A3"].font)
                sheet.cell(row, col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            except AttributeError:
                continue

    for offset, _source_col in enumerate(data_cols, start=5):
        moment = _norm(sheet.cell(10, offset).value)
        fill = PatternFill("solid", fgColor=_moment_fill(moment))
        for row in range(1, data_header_row + 1):
            sheet.cell(row, offset).fill = copy(fill)

    for col in range(1, max_col + 1):
        cell = sheet.cell(data_header_row, col)
        cell.fill = copy(data_header_fill if col <= 4 else sheet.cell(data_header_row - 1, col).fill)
        cell.font = copy(sheet["A11"].font)
        cell.alignment = copy(sheet["A11"].alignment)

    for row in range(data_header_row + 1, max_row + 1):
        for col in range(1, 5):
            sheet.cell(row, col).fill = copy(label_fill)
        for col in range(5, max_col + 1):
            sheet.cell(row, col).fill = copy(no_fill)

    _set_range_border(sheet, 1, 1, max_row, max_col, outside_style="thin")


def _moment_fill(moment: str) -> str:
    palette = {
        "a0": "FCE4D6",
        "a1": "BDD7EE",
        "a2": "E2EFDA",
        "a3": "FFF2CC",
        "a4": "FCE4D6",
        "a5": "BDD7EE",
        "h1": "DDEBF7",
    }
    return palette.get(moment, "DDEBF7")


def _max_numeric_value(values: Iterable[object]) -> int:
    numeric_values = []
    for value in values:
        if isinstance(value, (int, float)):
            numeric_values.append(int(value))
    return max(numeric_values, default=0)


def _count_treatments(sheet: Worksheet) -> int:
    table = _extract_treatment_table(sheet)
    treatment_numbers = []
    for row in table[2:]:
        value = row[0] if row else None
        if isinstance(value, (int, float)):
            treatment_numbers.append(int(value))
    return max(treatment_numbers, default=0)


def _last_non_empty_col(sheet: Worksheet, row: int) -> int:
    last = 0
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row, col).value not in (None, ""):
            last = col
    return last


def _last_non_empty_col_in_rows(sheet: Worksheet, rows: Iterable[int]) -> int:
    last = 1
    for row in rows:
        last = max(last, _last_non_empty_col(sheet, row))
    return last


def _evaluation_label(sheet: Worksheet, col: int) -> str:
    title = _cell_text(sheet.cell(12, col).value)
    unit = _cell_text(sheet.cell(15, col).value)
    class_name = _cell_text(sheet.cell(14, col).value)
    parts = [part for part in [title, class_name, unit] if part]
    return " - ".join(parts)


def _evaluation_day(value) -> str:
    text = _cell_text(value)
    return text if text else ""


def _unit_from_text(value) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    pieces = text.split()
    return pieces[-1] if pieces else text


def _clean_code(value) -> str:
    text = _cell_text(value)
    return re.sub(r"^\d+\s*;?\s*", "", text).strip()


def _has_brackets(value) -> bool:
    text = _cell_text(value)
    return "[" in text or "]" in text


def _extract_protocol_code(text: str) -> str:
    match = re.search(r"Protocolo:\s*([A-Z0-9_-]+)", text or "", flags=re.IGNORECASE)
    return match.group(1).strip() if match else ""


def _safe_title(text: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', " ", text or "").strip()
    return re.sub(r"\s+", " ", cleaned)[:90] or "Ensayo"


def _cell_text(value) -> str:
    return "" if value is None else str(value).replace("\xa0", " ").strip()


def _norm(value) -> str:
    text = _cell_text(value).lower()
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text



